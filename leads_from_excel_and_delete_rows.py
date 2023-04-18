import openpyxl
import requests
import json
#replace the data and url authentication
#for any help contact amjid@syncbricks.com
#This code is to create Leads from Excel file. The file path is provided. The format must be XLSX
#This will delete the row from excel once the lead is created in ERPNext
# Replace the URL, API key, and API secret with your own ERPNext instance's values
url = 'https://erp.syncbricks.com/api/resource/Lead'
api_key = 'asdf'
api_secret = 'asdf'

# Set the file path of the Excel sheet containing the leads data
leads_file_path = r"C:\Data\Dev\python\erpnext_api\files\leads.xlsx"

# Load the Excel sheet into memory
workbook = openpyxl.load_workbook(leads_file_path)
sheet = workbook.active

# Loop through each row of the sheet and add the lead data to ERPNext
for row_index in range(2, sheet.max_row + 1):
    # Create a dictionary with the data for the new lead
    new_lead = {
        'lead_name': sheet.cell(row=row_index, column=1).value,
        'lead_email': sheet.cell(row=row_index, column=2).value,
        'lead_phone': sheet.cell(row=row_index, column=3).value,
        'lead_status': sheet.cell(row=row_index, column=4).value
    }

    # Convert the dictionary to a JSON string
    json_data = json.dumps(new_lead)

    # Set the headers for the API request
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Token {api_key}:{api_secret}'
    }

    # Make the API request to insert the new lead
    response = requests.post(url, headers=headers, data=json_data)

    # Check the response status code
    if response.status_code == 200:
        print('New lead inserted successfully!')
        # Delete the row from the sheet
        sheet.delete_rows(row_index, 1)
        # Decrement the row index since we deleted a row
        row_index -= 1
    else:
        print('Error inserting new lead.')
        print(response.content)

# Save the changes to the Excel sheet
workbook.save(leads_file_path)
